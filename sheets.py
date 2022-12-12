import os
import re
from datetime import datetime
from openpyxl.styles import Font
import openpyxl

from employers import Employer


class Sheets:
    base_dir = os.getcwd()

    def book_activate(self, path):
        work_book = openpyxl.load_workbook(filename=path)
        return work_book


class InputSheet(Sheets):

    def get_path(self):
        file_name = None
        for f in os.listdir():
            if re.match(r'[Аа]рхив.(xlsx|xls)', f.lower()):
                file_name = f
                break
        if file_name:
            self.path_to_output = os.path.join(super().base_dir, file_name)
            return self.path_to_output
        else:
            return False

    def get_active_sheet(self):
        path = self.get_path()
        if path:
            self.book = super().book_activate(path)
        else:
            raise FileNotFoundError('Файл с именем Архив отсутствует в рабочей директории')
        self.book.active = 1
        sheet = self.book.active
        return sheet

    def get_employers(self, list):
        constructors = []
        designers = []

        for employer in list:
            constructors.append(Employer(employer['name'], employer['surname'], employer['stage']))

        return constructors

    def parse_input_sheet(self, list, year, month):
        self.c = self.get_employers(list)

        sheet = self.get_active_sheet()
        constructor_list = []

        for i, m in enumerate(self.c):
            constructor_list.append({'name': self.c[i].name,
                                     'surname': self.c[i].surname,
                                     'stage': self.c[i].stage,
                                     'initials': self.c[i].get_initials(),
                                     })
            constructor_list[i]['data'] = []
            for cel in range(1, 50000):
                if sheet.cell(row=cel, column=6).value:
                    d = sheet.cell(row=cel, column=5).value
                    if isinstance(d, datetime) and d.year == year and d.month == month:
                        if sheet.cell(row=cel, column=4).value:
                            point = sheet.cell(row=cel, column=4).value
                        else:
                            point = ''
                        d = d.strftime("%d-%m-%Y")
                        if re.findall(r'{}'.format(self.c[i].get_initials()), sheet.cell(row=cel, column=6).value):
                            constructor_list[i]['data'].append({'title': (sheet.cell(row=cel, column=2).value)[11:],
                                                                'date': d, 'point': point})

        return constructor_list


class OutputSheet(Sheets):
    def __init__(self, name, surname, stage, month, year):
        self.name = name
        self.surname = surname
        # self.data=data
        self.month = month
        self.year = year
        self.point_counter = 0
        self.x_cell = 5
        self.stage = stage

    def get_month(self):
        if isinstance(self.month, int) != True or self.month > 12:
            raise ValueError('Некорректный месяц')
        month_dict = {
            'Январь': 1,
            'Февраль': 2,
            'Март': 3,
            'Апрель': 4,
            'Май': 5,
            'Июнь': 6,
            'Июль': 7,
            'Август': 8,
            'Сентябрь': 9,
            'Октябрь': 10,
            'Ноябрь': 11,
            'Декабрь': 12,
        }
        for m, dig in month_dict.items():
            if self.month == dig:
                self.month_for_str = m

    def get_path(self):
        self.path_to_output = os.path.join(super().base_dir, '{} {}.xlsx'.format(self.name, self.surname))
        return self.path_to_output

    def del_sheet(self, sheet):
        for i in range(1, 500):
            for j in range(1, 5):
                if sheet.cell(row=i, column=j).value != None:
                    sheet.cell(row=i, column=j).value = None
                    sheet.cell(row=i, column=j).font = Font(bold=False)

    def output_sheet_activate(self):
        path = self.get_path()
        self.wb = openpyxl.load_workbook(filename=path)
        self.wb.active = 0
        self.sheet = self.wb.active
        return self.sheet

    # Создание заголовка
    def create_head(self, sheet):
        try:
            sheet.cell(row=1, column=1).value = 'Отчет {month} {year}г.'.format(month=self.month_for_str, year=self.year)
            sheet.cell(row=1, column=1).font = Font(bold=True, size=15)
            sheet.cell(row=2, column=1).value = '{0} {1}.'.format(self.surname, self.name)
            sheet.cell(row=2, column=1).font = Font(bold=True, size=15)
        except: raise ValueError(f'Ошибка при создании заголовка {self.name} {self.surname}')

    # Выбор работника из списка
    def create_data(self, list):
        self.employer_data = None
        for n, e in enumerate(list):
            if e['name'] == self.name and e['surname'] == self.surname:
                if not list[n]:
                    raise AttributeError(f'Нет такого работника {self.name} {self.surname}')
                self.employer_data = list[n]
                return self.employer_data

    # Построение списка с баллами
    def create_point_list(self, data, sheet, price):
        # Изменение ширины ячеек столбца А
        sheet.column_dimensions['A'].width = 65

        sheet.cell(row=4, column=1).value = 'Образцы'
        sheet.cell(row=4, column=1).font = Font(bold=True)
        sheet.cell(row=4, column=2).value = 'Баллы'
        sheet.cell(row=4, column=2).font = Font(bold=True)
        for d in data['data']:
            if re.findall(r'[Оо]браз', d['title']) or str(d['point']).isdigit():
                sheet.cell(row=self.x_cell, column=1).value = d['title']
                if d['point']:
                    sheet.cell(row=self.x_cell, column=2).value = d['point']
                    self.point_counter += int(d['point'])
                else:
                    sheet.cell(row=self.x_cell, column=2).value = 0
                self.x_cell += 1
                d['title'] = 0

        sheet.cell(row=self.x_cell, column=1).value = 'Сумма'
        sheet.cell(row=self.x_cell, column=1).font = Font(bold=True)
        sheet.cell(row=self.x_cell, column=2).value = self.point_counter
        sheet.cell(row=self.x_cell, column=2).font = Font(bold=True)
        sheet.cell(row=self.x_cell, column=3).value = self.point_counter * price
        sheet.cell(row=self.x_cell, column=3).font = Font(bold=True)
        self.x_cell += 1
        return data

    def get_non_point_list(self, data, sheet):
        self.x_cell += 2
        sheet.cell(row=self.x_cell, column=1).value = 'Задания'
        sheet.cell(row=self.x_cell, column=1).font = Font(bold=True)
        self.x_cell += 1
        for d in data['data']:
            if d['title']:
                sheet.cell(row=self.x_cell, column=1).value = d['title']
                self.x_cell += 1

    def get_designer_list(self, data, sheet):
        x = 4

        # Изменение ширины ячеек столбца
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 550
        sheet.column_dimensions['B'].width = 50

        for d in data['data']:
            sheet.cell(row=x, column=1).value = x - 3
            sheet.cell(row=x, column=2).value = d['title']
            sheet.cell(row=x, column=3).value = d['date']
            x += 1

    def wb_save(self):
        try:
            self.wb.save('{} {}.xlsx'.format(self.name, self.surname))
            self.wb.close()
        except: raise FileNotFoundError(f'Проверьте наличие файла {self.name} {self.surname}, либо закройте его')

    def creating(self, employer_dict):
        sheet = self.output_sheet_activate()
        self.get_month()
        self.del_sheet(sheet)
        self.create_head(sheet)
        self.data = self.create_data(employer_dict)

        if self.stage == 'Конструктор':
            d = self.create_point_list(self.data, sheet, 200)
            self.get_non_point_list(d, sheet)
        elif self.stage == 'Дизайнер':
            self.get_designer_list(self.data, sheet)
        else:
            d = self.create_point_list(self.data, sheet, 50)
            self.get_non_point_list(d, sheet)

        self.wb_save()
